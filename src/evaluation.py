import pandas as pd
import numpy as np
import re
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay
import sklearn
from datasets import load_metric
import pickle
import matplotlib.pyplot as plt
from openpyxl.styles.borders import Border, Side
import openpyxl
from openpyxl.styles import Alignment, Font

class Evaluate(object):
    """
    This class uses a saved model (a folder with .index and .vocabulary, .config)
    and its already computed predictions (model_test_predictions.xslx) to compute
    extracted labels, confusion matrixes and evaluation scores.
    
    The class purpose is to:
    
    
    
    
    """
    def __init__(self,model_path):
        """
        
        """
        super().__init__()
        
        # Setting paths for evaluation------------------------------------------
        self.model_path=model_path # PAth to the model folder
        # Path to the .config file inside the model folder
        self.config_path=model_path+'/'+model_path.split('/')[-1]+'.config'
        # Path where to save the predictions 
        self.predictions_path=model_path+'/'+model_path.split('/')[-1]+'_test_predictions.xlsx'
        # Path were to save the labels that will be extracted from the 
        # predictions.
        self.labels_path=model_path+'/'+model_path.split('/')[-1]+'_test_labels.csv'
        # Path for saving confusion matrixes. 
        self.fig_path=model_path+'/'+model_path.split('/')[-1]+'_cm_xx.png'
        # Path where to save the test metrics.
        self.scores_path=model_path+'/'+model_path.split('/')[-1]+'_test_metrics.csv'
        # The df with all the predictions. 
        self.predictions=pd.read_excel(self.predictions_path)
        # We initialize the scores dictionary
        self.scores={}
         
        # Retrieve configuration dictionary-------------------------------------
        with open(self.config_path, 'rb') as fp:
            self.config = pickle.load(fp) 
        
    def get_label(self,caption,label):
        """Extracts the label indicated by the string 'label'.
        from the caption. 
        
        This function is to be used as a map function from within a dataframe
        call. Ex: df.true_captions.map(lambda x: get_label(x,label))
        
        Args:
            caption(series_row): A row containing a caption. 
            label(str):
        """
        rgx={'core_modality':r'core modality:([ a-z/0-9]+)modality',
             'plane': r'plane:([ a-z/0-9]+)anatomy',
             'anatomy':r'anatomy:([ a-z/0-9]+)findings'}
        caption=caption.lower() # lowercase
        # find regex match
        match=re.search(rgx[label],caption)
        if match is None:
        # If no match is found return 'N/A'
            return 'N/A'    
        label=match.group(1) 
        label=label.strip() # strip white spaces
        return label
    
    def plot_conf_matrix(self,y_true,y_pred,path):
        """Plots and saves the confusion matrix."""
        uniques = list(set(np.concatenate([y_true,y_pred])))
        cm = confusion_matrix(y_true, y_pred, labels=uniques)
        disp = ConfusionMatrixDisplay(confusion_matrix=cm,
                                  display_labels=uniques)

        fig=disp.plot(xticks_rotation='vertical').figure_
        fig.savefig(path, dpi='figure', format=None, metadata=None,
            bbox_inches='tight', pad_inches=0.3,
            facecolor='auto', edgecolor='auto',
            backend=None)
        plt.close('all')
        
    def get_labels(self):
        """Get labels for all the predictions."""
        df=self.predictions
        # Extract true modality, plane and anatomy labels--------------------------------------- 
        # extract true core modality
        df['true_core_modality']=df['true_caption_1'].apply(lambda c:self.get_label(c,'core_modality'))
        # extract true plane
        df['true_plane']=df['true_caption_1'].apply(lambda c:self.get_label(c,'plane')) 
        # extract true anatomy
        df['true_anatomy']=df['true_caption_1'].apply(lambda c:self.get_label(c,'anatomy')) 

        # Extract predicted modality, plane and anatomy labels--------------------------------------- 
        # extract true core modality
        df['predicted_core_modality']=df['predicted_caption'].apply(lambda c:self.get_label(c,'core_modality'))
        # extract true plane
        df['predicted_plane']=df['predicted_caption'].apply(lambda c:self.get_label(c,'plane')) 
        # extract true anatomy
        df['predicted_anatomy']=df['predicted_caption'].apply(lambda c:self.get_label(c,'anatomy'))
        self.predictions=df

        df.to_csv(self.labels_path)
        
    def confusion_matrixes(self):
        """Create all confusion matrixes. """
        # Extract labels and place them into columns
        self.get_labels() 
        
        # Plot and save confusion matrixes.
        df=self.predictions
        self.plot_conf_matrix(y_true=df['true_core_modality'],
                         y_pred=df['predicted_core_modality'],
                         path=self.fig_path.replace('xx','cm'))

        self.plot_conf_matrix(y_true=df['true_plane'],
                         y_pred=df['predicted_plane'],
                         path=self.fig_path.replace('xx','pl'))

        self.plot_conf_matrix(y_true=df['true_anatomy'],
                         y_pred=df['predicted_anatomy'],
                         path=self.fig_path.replace('xx','an'))
        
    def classification(self):
        """Obtain f1 weighted score for core modality, plane and anatomy"""
        df=self.predictions
        y_true=df['true_core_modality']
        y_pred=df['predicted_core_modality']
        f1w_cm=sklearn.metrics.f1_score(y_true, y_pred,average='weighted')

        # Plane f1 weighted
        y_true=df['true_plane']
        y_pred=df['predicted_plane']
        f1w_pl=sklearn.metrics.f1_score(y_true, y_pred,average='weighted')

        # Anatomy f1 weighted
        y_true=df['true_anatomy']
        y_pred=df['predicted_anatomy']
        f1w_an=sklearn.metrics.f1_score(y_true, y_pred,average='weighted') 
        
        # Record classification scores to scores dictionary. 
        self.scores['core_modality_f1_w']=f1w_cm
        self.scores['plane_f1_w']=f1w_pl
        self.scores['anatomy_f1_w']=f1w_an
        
    def extract_true_captions_list(self,row,references,true_caption_cols):
        """To be used as map function through df.apply() in a 
        results df obtain from a results.csv file. 
        It iterates over rows of the df and makes 
        a list of lists with all reference sentences in a single row. 
        The resulting list is the "references" parameter. 
        """
        ref_row=[] # make a list of true captions for a single row
        for col in true_caption_cols:
        # for every column that contains true captions.
            ref_row.append(row[col]) # Append it to this row's list
        references.append(ref_row) # Append the row list of captions to records. 
        return 
    
    def bleu(self):
        df=self.predictions
        # Obtain titles of columns refering to true captions. 
        true_caption_cols=[]
        for col_name in list(df.columns):
            if 'true_caption' in col_name:
                true_caption_cols.append(col_name)
        # Create empty references list. 
        references=[]
        # Fill references list.
        df.apply(lambda row:self.extract_true_captions_list(row,references,true_caption_cols),axis=1)
        
        self.references_list=references
        predictions=df['predicted_caption'].to_list() # Fill predictions list
        self.predictions_list=predictions
        # compute sacrebleu metric
        sacrebleu = load_metric('sacrebleu')
        sacrebleu_score=sacrebleu.compute(predictions=predictions, references=references)
        self.scores['bleu'] = sacrebleu_score['score']
        self.scores['mp1'] = sacrebleu_score['precisions'][0]
        self.scores['mp2'] = sacrebleu_score['precisions'][1]
        self.scores['mp3'] = sacrebleu_score['precisions'][2]
        self.scores['mp4'] = sacrebleu_score['precisions'][3]
        self.scores['bp'] = sacrebleu_score['bp']
        self.save_scores()
    
    def rouge(self):
        rouge=load_metric('rouge')
        rouge_score=rouge.compute(predictions=self.predictions_list,
                                  references=self.references_list)
        self.scores['rouge1_f_low']=rouge_score['rouge1'].low.fmeasure
        self.scores['rouge1_f_mid']=rouge_score['rouge1'].mid.fmeasure
        self.scores['rouge1_f_high']=rouge_score['rouge1'].high.fmeasure

        self.scores['rouge2_f_low']=rouge_score['rouge2'].low.fmeasure
        self.scores['rouge2_f_mid']=rouge_score['rouge2'].mid.fmeasure
        self.scores['rouge2_f_high']=rouge_score['rouge2'].high.fmeasure

        self.scores['rougeL_f_low']=rouge_score['rougeL'].low.fmeasure
        self.scores['rougeL_f_mid']=rouge_score['rougeL'].mid.fmeasure
        self.scores['rougeL_f_high']=rouge_score['rougeL'].high.fmeasure
        self.save_scores()
        
    def save_scores(self):
        
        # Convert values to lists so that they are compatible with pd.DataFrame
        df_dict={}
        for key,value in self.config.items():
            df_dict[key]=[value]
     
        # Add the metrics to the dictionary and save it to csv
        for key,value in self.scores.items():
            df_dict[key]=[value]
            pd.DataFrame(df_dict).to_csv(self.scores_path,index=False)
            
        self.results=df_dict
        
    def evaluate_all(self):
        """"Run evaluation metrics and save results."""
        if self.config['dataset']=='medpix':
            print('Making confusion matrixes.')
            self.confusion_matrixes()
            self.classification()
        print('Calculating global bleu scores.')
        self.bleu()
        print('Calculating global rouge scores.')
        self.rouge()
        print('Calculating individual bleu and rouge scores.')
        self.evaluate_individual_predictions()
    
    def evaluate_individual_predictions(self):
        """ Adds  bleu and rouge columns with scores for each perediction in the 
        predictions excel."""
        df=self.predictions
        reference_cols=[col for col in df.columns.to_list() if 'true' in col]
        bleu=load_metric('sacrebleu')
        rouge=load_metric('rouge')
        bleus=[]
        rouges=[]
        # Compute metrics for each row in the dataframe
        for index, row in df.iterrows():
            predictions=[row['predicted_caption']]
            references=[row[reference_cols].to_list()]
            # compute metrics
            bleu_dict=bleu.compute(predictions=predictions,
                                   references=references)
            rouge_dict=rouge.compute(predictions=predictions,
                                   references=references)
            # append relevant scores
            bleus.append(bleu_dict['score'])
            # rouges.append
            rouges.append(rouge_dict['rougeL'].mid.fmeasure)

        wb = openpyxl.load_workbook(self.predictions_path) # open excel file
        ws = wb.active # Select the active worksheet

        # Get the letter of the column after last filled column and the next one
        col_letter_bleu = openpyxl.utils.cell.get_column_letter(ws.max_column+1)
        col_letter_rouge = openpyxl.utils.cell.get_column_letter(ws.max_column+2)
        # Add the column title for bleu
        ws[col_letter_bleu+str(1)]='bleu' # Set column title
        ws[col_letter_bleu+str(1)].font = Font(bold=True) # Set bold font
        ws[col_letter_bleu+str(1)].alignment = Alignment(horizontal='center') # Center text
        # Add borders
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
        ws[col_letter_bleu+str(1)].border = thin_border

        # Add column title for rouge
        ws[col_letter_rouge+str(1)]='rougeL_f_mid' # Set column title
        ws[col_letter_rouge+str(1)].font = Font(bold=True) # Set bold font
        ws[col_letter_rouge+str(1)].alignment = Alignment(horizontal='center') # Center text
        # Add borders
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
        ws[col_letter_rouge+str(1)].border = thin_border

        # Add scores to the bleu column iteratively
        for i,score in enumerate(bleus):
        # For every score
            # We add two to the row because excel index starts at 1 and we must skip
            # the column titles row.
            row = i+2 
            cell_index = col_letter_bleu+str(row) # Get index in LetterNumber format
            ws[cell_index]=score # Write score to cell

        # Add scores to the rouge column iteratively
        for i,score in enumerate(rouges):
        # For every score
            # We add two to the row because excel index starts at 1 and we must skip
            # the column titles row.
            row = i+2 
            cell_index = col_letter_rouge+str(row) # Get index in LetterNumber format
            ws[cell_index]=score # Write score to cell

        wb.save(self.predictions_path)
        
        