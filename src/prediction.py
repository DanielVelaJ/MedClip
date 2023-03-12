"""Model to predict captions using a trained captioning transformer.

This module contains the 'Predict' class which contains methods that are used
along a CaptioningTransformer and a CustomTokenizer (or subclasses of 
CustomTokenizer) to generate captions over a given dataset. 

By Daniel Vela (d.vela1@hotmail.com)

"""
import tensorflow as tf
import re
import numpy as np
import matplotlib.pyplot as plt
import time
import datetime
import textwrap
import json
import pandas as pd
import os
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Alignment, Font
from PIL import Image
import shutil
from openpyxl.styles.borders import Border, Side

class Predict(object):
    """ This class uses a model and a dataset to create predictions.
    
    The class makes predictions using a model on a test dataset. 
    The test dataset must contain the "image" key, containing (299,299,3) images
    and a "tokens" key contianing (num_capts_per_img,tokens) tokens for each 
    image. This dataset must be a tensorflow dataset like those created using 
    the pipeline.py module. After generating the predicted captions using the 
    model, the class saves an excel file that contains the ture and predicted 
    captions as well as a column with the images used to produce them. 
    
    
    """
    def __init__(self, dataset, model_path):
        """
        Args:
            dataset(tf dataset): A tensorflow dataset containing images 
                and texts. 
            model_path(str): The path to the folder where the experiment lives 
                and where all predictions will be saved. 
        """
        self.dataset=dataset
        
        # The path to the predictions excel that will contain all predictions. 
        self.predictions_path=(model_path
                               +'/'
                               +model_path.split('/')[-1]
                               +'_test_predictions.xlsx')
        self.predictions_csv_path=(model_path
                               +'/'
                               +model_path.split('/')[-1]
                               +'_test_predictions.csv')
       
        
        
    
    @classmethod
    def image_caption(cls, img,model,tokenizer,show=False):
        """ Uses the given image, model and tokenizer to 
        produce a caption. 
        Args: 
            img: (299,299,3) array of floats. 
            model: a CaptioningTransformer.
            tokenizer: A CustomTokenizer son.
            show(optional bool): Whether to print results. 

        Returns
            pred_caption (str): predicted caption.


        """
        img=tf.expand_dims(img,0) # Add batch dimension to image. 
        # Initializing output to <start> token
        pred_tokens=tokenizer.tokenize('<start>')[0:-1]
        # Adding batch dimension and converting to numpy
        pred_tokens=tf.expand_dims(pred_tokens,0).numpy()
        for i in range(model.config['seq_len']-2):
        # For every prediction step in the sequence:
            mask=tf.math.not_equal(pred_tokens,0) # get mask
            probs=model.calculate_probs(img,pred_tokens,mask) # obtain next word probs
            # obtain the vocab index/token of the next highest probable word for the 
            # current prediction step [i]
            prediction=tf.argmax(probs,-1)[0][i].numpy() 
            if prediction==tokenizer.tokenize('<end>').numpy()[0]:
            # If the prediction matches the end token stop the process. 
                break
            pred_tokens[0,i+1]=prediction # Append the predicted token to the output
        # convert predicted token sequence to words. 
        pred_caption=tokenizer.untokenize(pred_tokens[0][1:])
        if show==True:
            plt.imshow(img.numpy()[0].astype(int))
            plt.show()
            print('AI-Generated:')
            print(pred_caption)

        return pred_caption
    
    @classmethod
    def dataset_captions_individually(cls, 
                                      dataset,
                                      model,
                                      tokenizer,
                                      model_path,show=True):
        """
        Predict captions for tha whole dataset and save them. 
        
        Notice that captions are predicted image by image in contrast to 
        'dataset_captions' which predicts captions batch by batch.
        
        Args:
            dataset(pipeline): A dataset pipeline built by 
                pipelines.buil_pipeline()
            model(CaptioningTransformer): a captioning transformer model.
            tokenizer(CustomTokenizer): a tokenizer object.
            model_path(str): Path to model folder in which predictions will 
                be saved
            show(bool,optional): Whether to print images and captions, defaults
                to True. 
            
        
        """
        d={'predicted_caption':[]}
        for i,s in tqdm(enumerate(dataset)):
            img=s['image'] # Obtain image
            # predict caption and save it to dictionary.
            d['predicted_caption'].append(cls.image_caption(img,model,
                                              tokenizer,
                                              show=show))
            # Get the true token sequences there may be more 
            # than one per image. 
            true_tokens=s['tokens']
            for j in range(true_tokens.shape[0]):
            # For every caption in the group of captions for the same image.
                # Untokenize the caption while excluding the start token
                caption=tokenizer.untokenize(true_tokens[j][1:])
                # Eliminate the <end> word in the untokenized caption.
                # and strip white spaces.
                caption=caption[0:caption.index('<end>')].strip()
                print('\nTrue caption(s):')
                print(caption)
                # Save caption to dictionary. 
                col_name='true_caption_'+str(j+1)
                if i==0:
                # If this is the first time the key is used create list.
                    d[col_name]=[caption]
                else: 
                # Otherwise append to list. 
                    d[col_name].append(caption)

            # results=pd.DataFrame(d)
            # os.makedirs(model_path,exist_ok=True)
            # predictions_path=(model_path
            #                   +'/'
            #                   +model_path.split('/')[-1]
            #                   +'_test_predictions.csv')
            # results.to_csv(predictions_path,index=False)
            
           
    def dataset_captions(self, 
                         model,
                         tokenizer,
                         batch_size=32):
        """
        Predict captions for tha whole dataset and save them. 
        
        Args:
            model(CaptioningTransformer): a captioning transformer model.
            tokenizer(CustomTokenizer): a tokenizer object.
            batch_size(str, optional): The batch size for every prediction step.
                captions will be predicted batch by batch. 
            
        
        """
        dataset=self.dataset # obtain the dataset from instance property. 
        
        results_dict={} # Dictionary to store final df
        # Make caption predictions using the model--------------------------------------
        results_dict['predicted_caption']=[] # Initialize list to hold captions
        end_token=tokenizer.tokenize('<end>')[0].numpy() # Get end token
        for batch in tqdm(dataset.batch(batch_size)):
        # For every batch in the dataset
            imgs=batch['image'].numpy() # Get the batch of images
            # Initialize an array of start tokens of size 
            # (batch_size,seq_len)
            pred_tokens=tokenizer.tokenize(['<start>']*imgs.shape[0])[:,0:-1].numpy()

            # Begin the prediction process
            for step in range(model.config['seq_len']-2):
            # For every prediction step:
                mask=tf.not_equal(pred_tokens,0)# obtain masks
                # obtain next word probabilities for batch
                # these will have shape (batch_size,seq_len,vocab_Size)
                probs=model.calculate_probs(imgs,pred_tokens,mask)
                # Obtain predicted tokens for this step for all the batch
                # (batch_size,1)
                step_tokens=tf.argmax(probs,-1)[:,step]
                # Assign calculated tokens to the prediction sequences along the 
                # sequence length dimension. We do this for all batches at once.
                pred_tokens[:,step+1]=step_tokens
                if np.all(np.any(pred_tokens==end_token,axis=-1)):
                # If all of the sequences have reached an end token, we exit the 
                # prediction loop
                    break

            # Convert the predicted token sequences in the batch to captions.
            for sequence in pred_tokens:
            # For every sequence in the batch:
                caption=tokenizer.untokenize(sequence) # untokenize the sequence
                # remove <start> and <end> tokens
                caption=caption[7:caption.find('<end>')].strip() 
                results_dict['predicted_caption'].append(caption) # Append to dict list


        # Obtain the true captions------------------------------------------------------
        # Initialize empty lists for every column in the dictionary that will hold
        # the true captions. There is one column for every caption that corresponds
        # to the same image. 
        for n in range(batch['tokens'].shape[1]):
            col_name='true_caption_'+str(n+1) 
            results_dict[col_name]=[]


        for row in dataset:
        # for every row in the dataset (captions_per_image,seq_len)
        # Get the token sequences that correspond to the same image
            token_sequences=row['tokens'] 
            for n,token_sequence in enumerate(token_sequences):
            # for every sequence in the group of sequences (seq_len),
                # starting from the second token in the sequence to skip the 
                # <start> token:
                caption=tokenizer.untokenize(token_sequence) # Untokenize sequence
                # Get rid of  <start> and <end> tokens
                caption=caption[7:caption.find('<end>')].strip() 
                # define the column name to save the caption. There is one column for
                # every caption in 'captions_per_image'
                col_name='true_caption_'+str(n+1) 
                results_dict[col_name].append(caption)

        # Create a dataframe from the results dictionary and save as csv.
        pd.DataFrame(results_dict).to_excel(self.predictions_path,index=False)
        pd.DataFrame(results_dict).to_csv(self.predictions_csv_path,index=False)
        
    
    def append_images(self):
        """ Adds a column to the predictions excel file with images. 
        
        The excel file is first formatted to recieve images and then 
        the images corresponding ot each prediction are appended in the "image"
        column of the predictions excel. 
        
        """
        # Make a temporal folder and then loop through all images in the dataset
        # to save them in this temporal folder as png. 
        os.makedirs('temp_images',exist_ok=True) # make a temporal image folder
        for i,sample in tqdm(enumerate(self.dataset)):
        # For every sample in the dataset:
            # Retrieve image array in uint8
            img=sample['image'].numpy().astype('uint8') 
            img = Image.fromarray(img, "RGB") # Convert to image format.
            newsize = (100, 100)
            img = img.resize(newsize)
            img.save('temp_images/image'+str(i)+'.png') # Save the image as png.


        # Add an empty column with title 'image' to the excel
        wb = openpyxl.load_workbook(self.predictions_path) # Open excel file
        ws = wb.active # Select the active worksheet
        ws.insert_cols(1) # Add column at start of file.
        ws['A1']='image' # Set column title
        ws['A1'].font = Font(bold=True) # Set bold font
        ws['A1'].alignment = Alignment(horizontal='center') # Center text
        ws.column_dimensions['A'].width = 20 # Set column width
        # Add borders
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
        ws['A1'].border = thin_border
  
        # Apply formatting to cells containing predictions and true captions. 
        # Set width for rows containing captions
        for col_idx in range(2,ws.max_column+1):
            col_letter = openpyxl.utils.cell.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 70

        # Setup cells wrapup property for cells containing captions
        for row in ws.iter_cols(min_row=2, min_col=2, 
                                max_row=ws.max_row, max_col=ws.max_column):  
            for cell in row:  
                a=cell
                cell.alignment = Alignment(wrap_text=True,vertical='top') 

        # Loop over temporaly saved images and add them to the excel
        for i,sample in tqdm(enumerate(self.dataset)):
        # For every image in the dataset:
            # Set the row height, we add 2 to i because rows in excel start at 1
            # and we must skip the column titles row. 
            ws.row_dimensions[i+2].height = 100
            # Add the image 
            img = openpyxl.drawing.image.Image('temp_images/image'+str(i)+'.png')
            anchor = 'A'+str(i+2) # Coordinates where image will be inserted
            img.anchor = anchor
            img.width = 130
            img.height = 130
            ws.add_image(img)
        wb.save(self.predictions_path) # save excel
        # remove temporal images folder
        # remove image png 
        shutil.rmtree('temp_images', ignore_errors=False, onerror=None) 

    